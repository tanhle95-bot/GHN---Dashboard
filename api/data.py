from http.server import BaseHTTPRequestHandler
import json
import urllib.request
import io
import openpyxl

URL = "https://docs.google.com/spreadsheets/d/1Kjqwt6MIghCzfCSifVrpIpVHbC0o77lxMVVFlCZ26xY/export?format=xlsx"

class handler(BaseHTTPRequestHandler):
    def do_GET(self):
        try:
            # Tải file Excel về bộ nhớ
            req = urllib.request.Request(URL, headers={"User-Agent": "Mozilla/5.0"})
            response = urllib.request.urlopen(req)
            excel_data = io.BytesIO(response.read())
            
            # Khởi tạo workbook dạng đọc tối ưu (read_only và data_only)
            wb = openpyxl.load_workbook(excel_data, read_only=True, data_only=True)
            
            data = {
                "gtc_ratio": [],
                "delayed_orders": [],
                "b2b_orders": [],
                "installation_orders": {
                    "success": [],
                    "pending": [],
                    "discrepancy": []
                }
            }
            
            # 1. Xử lý GTC Ratio (Sheet: raw_hieusuat)
            if "raw_hieusuat" in wb.sheetnames:
                sheet = wb["raw_hieusuat"]
                rows = sheet.iter_rows(values_only=True)
                header = next(rows)
                
                col_wh = header.index("warehouse_name")
                col_sld = header.index("sld")
                col_sld_gtc = header.index("sld_gtc")
                
                hanoi_keywords = ["Hà Nội", "HNO", "Long Biên", "Bắc Từ Liêm", "Thanh Oai", "Hoài Đức", "Đức Long", "Thanh Trì", "Đông Anh"]
                
                groups = {}
                for row in rows:
                    if not row or len(row) <= max(col_wh, col_sld, col_sld_gtc):
                        continue
                    wh_val = str(row[col_wh]) if row[col_wh] is not None else ""
                    if "Kho Giao Hàng Nặng" in wh_val and any(kw in wh_val for kw in hanoi_keywords):
                        sld_val = float(row[col_sld]) if row[col_sld] is not None else 0.0
                        sld_gtc_val = float(row[col_sld_gtc]) if row[col_sld_gtc] is not None else 0.0
                        
                        if wh_val not in groups:
                            groups[wh_val] = {"total_assigned": 0.0, "total_success": 0.0}
                        groups[wh_val]["total_assigned"] += sld_val
                        groups[wh_val]["total_success"] += sld_gtc_val
                
                gtc_ratio_list = []
                for wh_name, vals in groups.items():
                    total_assigned = vals["total_assigned"]
                    total_success = vals["total_success"]
                    gtc_ratio = (total_success / total_assigned * 100) if total_assigned > 0 else 0.0
                    gtc_ratio_list.append({
                        "warehouse_name": wh_name,
                        "total_assigned": total_assigned,
                        "total_success": total_success,
                        "gtc_ratio": gtc_ratio
                    })
                
                gtc_ratio_list.sort(key=lambda x: x["gtc_ratio"], reverse=True)
                data["gtc_ratio"] = gtc_ratio_list
            
            # 2. Xử lý Đơn hoãn giao >3D (Sheet: 4. Đơn >3D)
            if "4. Đơn >3D" in wb.sheetnames:
                sheet = wb["4. Đơn >3D"]
                rows = sheet.iter_rows(values_only=True)
                header = next(rows)
                
                delayed_keywords = ["hà nội", "hno", "long biên", "bắc từ liêm", "thanh oai", "hoài đức", "đức long", "thanh trì", "đông anh"]
                
                delayed_list = []
                for row in rows:
                    if not row or len(row) < 5:
                        continue
                    wh_val = row[1]
                    if wh_val is not None:
                        wh_str = str(wh_val).lower()
                        if "kho giao hàng nặng" in wh_str and any(kw in wh_str for kw in delayed_keywords):
                            total_3d = float(row[2]) if row[2] is not None else 0.0
                            over_7d = float(row[3]) if row[3] is not None else 0.0
                            four_to_seven = float(row[4]) if row[4] is not None else 0.0
                            
                            if over_7d >= 10 or total_3d >= 100:
                                status = "Cảnh báo"
                            elif over_7d > 0 or total_3d >= 20:
                                status = "Cần lưu ý"
                            else:
                                status = "Bình thường"
                                
                            delayed_list.append({
                                "warehouse": str(wh_val),
                                "total_3d": total_3d,
                                "over_7d": over_7d,
                                "4_to_7d": four_to_seven,
                                "status": status
                            })
                
                delayed_list.sort(key=lambda x: x["total_3d"], reverse=True)
                data["delayed_orders"] = delayed_list
            
            # 3. Xử lý Đơn hàng B2B Ưu tiên (Sheet: 6.2 B2B | Đơn ƯU TIÊN GIAO)
            if "6.2 B2B | Đơn ƯU TIÊN GIAO" in wb.sheetnames:
                sheet = wb["6.2 B2B | Đơn ƯU TIÊN GIAO"]
                rows = sheet.iter_rows(values_only=True)
                header = next(rows)
                
                b2b_groups = {}
                priorities_set = set()
                hanoi_keywords_b2b = ["hà nội", "hno", "long biên", "bắc từ liêm", "thanh oai", "hoài đức", "đức long", "thanh trì", "đông anh"]
                
                for row in rows:
                    if not row or len(row) < 3:
                        continue
                    
                    has_hanoi = False
                    for val in row:
                        if val is not None:
                            val_str = str(val).lower()
                            if any(kw in val_str for kw in hanoi_keywords_b2b):
                                has_hanoi = True
                                break
                                
                    if has_hanoi:
                        priority_val = str(row[0]) if row[0] is not None else "Không xác định"
                        wh_val = str(row[2]) if row[2] is not None else "Không xác định"
                        
                        priorities_set.add(priority_val)
                        if wh_val not in b2b_groups:
                            b2b_groups[wh_val] = {}
                        if priority_val not in b2b_groups[wh_val]:
                            b2b_groups[wh_val][priority_val] = 0
                        b2b_groups[wh_val][priority_val] += 1
                
                b2b_orders_list = []
                for wh_name, prio_counts in b2b_groups.items():
                    record = {"warehouse": wh_name}
                    for prio in priorities_set:
                        record[prio] = prio_counts.get(prio, 0)
                    b2b_orders_list.append(record)
                
                data["b2b_orders"] = b2b_orders_list
            
            # 4. Xử lý Đơn lắp đặt (Sheet: 5. DS đơn Lắp đặt)
            if "5. DS đơn Lắp đặt" in wb.sheetnames:
                sheet = wb["5. DS đơn Lắp đặt"]
                rows = sheet.iter_rows(values_only=True)
                header = next(rows)
                
                col_wh_inst = header.index("Kho hiện tại")
                col_code_inst = header.index("Mã đơn hàng")
                col_status_inst = header.index("Trạng thái đơn hàng")
                col_type_inst = header.index("Type lắp đặt")
                col_buyer_inst = header.index("Buyer đồng ý lắp đặt?")
                
                for row in rows:
                    if not row or len(row) <= max(col_wh_inst, col_code_inst, col_status_inst, col_type_inst, col_buyer_inst):
                        continue
                    wh_val = str(row[col_wh_inst]) if row[col_wh_inst] is not None else ""
                    if "Hà Nội" in wh_val or "HNO" in wh_val:
                        type_inst = str(row[col_type_inst]) if row[col_type_inst] is not None else ""
                        buyer_inst = str(row[col_buyer_inst]) if row[col_buyer_inst] is not None else ""
                        
                        item = {
                            "order_code": str(row[col_code_inst]) if row[col_code_inst] is not None else "",
                            "warehouse": wh_val,
                            "status": str(row[col_status_inst]) if row[col_status_inst] is not None else "",
                            "install_type": type_inst
                        }
                        
                        if type_inst == "Đã lắp đặt":
                            data["installation_orders"]["success"].append(item)
                        elif buyer_inst == "Đồng ý" and type_inst != "Đã lắp đặt":
                            if type_inst == "Không cần lắp đặt":
                                data["installation_orders"]["discrepancy"].append(item)
                            else:
                                data["installation_orders"]["pending"].append(item)
            
            # Đóng workbook để giải phóng bộ nhớ
            wb.close()
            
            self.send_response(200)
            self.send_header('Content-type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps(data).encode('utf-8'))
            
        except Exception as e:
            self.send_response(500)
            self.send_header('Content-type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            error_data = {"error": str(e)}
            self.wfile.write(json.dumps(error_data).encode('utf-8'))
